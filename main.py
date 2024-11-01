import logging
import os
from datetime import datetime as dt
from multiprocessing.dummy import Pool as ThreadPool

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


BASE_URL = "https://yacht-parts.ru"
CATALOG_URL = "https://yacht-parts.ru/catalog/"
BRANDS_URLS = [
    "https://yacht-parts.ru/info/brands/",
    "https://yacht-parts.ru/info/brands/?PAGEN_1=2",
    "https://yacht-parts.ru/info/brands/?PAGEN_1=3",
]

OUTPUT_FOLDER = "output"
LOGS_FOLDER = "logs"


# Парсинг страницы каталога для получения всех категорий товаров
def parse_catalog() -> dict[str, str]:
    result = {}
    try:
        resp = requests.get(CATALOG_URL)
        soup = BeautifulSoup(resp.text, "html.parser")

        sections = soup.find_all("div", "section_item")
        for item in sections:
            category_name = item.find("li").a.span.text
            result[category_name] = [item.a.get('href') for item in item.find_all("li", "sect")]
        logging.info("Каталог успешно получен")
    except Exception as e:
        logging.exception(e)
    return result


# Парсинг каждой категории для получения кол-ва страниц пагинатора
def parse_page_numbers(url: str) -> tuple[str, int]:
    num = 1
    try:
        resp = requests.get(BASE_URL + url)
        soup = BeautifulSoup(resp.text, "html.parser")
        try:
            num = int([item for item in soup.find("span", "nums")][-2].text)
        except TypeError:
            num = 1
        logging.info(f"Получил кол-во страниц {resp.url}")

    except Exception as e:
        logging.exception(e)

    return (url, num)


# Парсинг адресов предметов на странице
def parse_page(data: tuple[str, int]) -> list[str]:
    items = []
    try:
        resp = requests.get(BASE_URL + data[0], params={"PAGEN_1": data[1]})
        soup = BeautifulSoup(resp.text, "html.parser")
        items = [item.a.get('href') for item in soup.find_all("div", "item-title")]
        logging.info(f"Page parsed: {resp.url}")

    except Exception as e:
        logging.exception(e)

    return items


# Парсинг предмета
def parse_item(url: str) -> dict[str, str]:
    result = {
        "title": "",
        "price": "",
        "article": "",
        "description": "",
        "images": [],
    }
    try:
        resp = requests.get(BASE_URL + url)
        soup = BeautifulSoup(resp.text, "html.parser")

        title = soup.find("h1", id="pagetitle").text
        price = soup.find("div", "price").text.strip()
        article = soup.find("div", "article iblock").find("span", "value").text
        description = soup.find("div", "preview_text").text.strip()
        images = soup.find("div", "slides").find_all("img")
        logging.info(f"Item parsed: {resp.url}")
        result = {
            "title": title,
            "price": price,
            "article": article,
            "description": description,
            "images": [image.get("src") for image in images],
        }

    except Exception as e:
        logging.exception(e)

    return result


# Парсинг существующих брендов
def parse_brands(url: str) -> list[str]:
    brand_list = []
    try: 
        resp = requests.get(url)
        soup = BeautifulSoup(resp.text, "html.parser")
        brand_list = list(map(
            lambda item: item["title"], 
            [item.find("img") for item in soup.find("ul", "brands_list")][1:-1:2]
        ))
        logging.info("Бренды успешно получены")
    except Exception as e:
        logging.exception(e)

    return brand_list


# Сохранение в эксель
def parse_items_and_save(filename: str, catalog: dict[str, str], brands: list[str]):
    workbook = Workbook() 

    for key in catalog:
        logging.info(f"Обработка: {key}")
        urls_with_nums = pool.map(parse_page_numbers, catalog[key])
        workbook_name = key[:31] # Название листа должно содержать < 31 символа
        workbook.create_sheet(workbook_name)
        sheet = workbook[workbook_name]
        items = []
        for url, num in urls_with_nums:
            all_urls = pool.map(parse_page, [(url, i) for i in range(1, num+1)])
            for page in all_urls:
                items += pool.map(parse_item, [item for item in page])

        sheet.cell(1, 1, "Название")
        sheet.cell(1, 2, "Бренд")
        sheet.cell(1, 3, "Категория")
        sheet.cell(1, 4, "Цена")
        sheet.cell(1, 5, "Артикул")
        sheet.cell(1, 6, "Описание")
        sheet.cell(1, 7, "Изображения")
        for i, item in enumerate(items, 2):
            item_brand = "Нет"
            for brand in brands:
                if brand in item.get("title", ""):
                    item_brand = brand

            sheet.cell(i, 1, item.get("title"))
            sheet.cell(i, 2, item_brand)
            sheet.cell(i, 3, key)
            sheet.cell(i, 4, item.get("price"))
            sheet.cell(i, 5, item.get("article"))
            sheet.cell(i, 6, item.get("description"))
            sheet.cell(i, 7, ', '.join(item.get("images", [])))
        workbook.save(f"{OUTPUT_FOLDER}/{filename}")

    del workbook["Sheet"]

    workbook.save(f"{OUTPUT_FOLDER}/{filename}")


if __name__== '__main__':
    if not os.path.isdir(OUTPUT_FOLDER):
        os.mkdir(OUTPUT_FOLDER)
    if not os.path.isdir(LOGS_FOLDER):
        os.mkdir(LOGS_FOLDER)

    start_time = dt.now()
    pool = ThreadPool(8)
    logging.basicConfig(
        level=logging.INFO,
        format="{asctime} - {levelname} - {message}", 
        filename=f"{LOGS_FOLDER}/app.log",
        style="{",
        filemode="a")

    logging.info("PROGRAM STARTED")

    brands = pool.map(parse_brands, BRANDS_URLS)

    brands = brands[0] + brands[1] + brands[2] # create 1 arr from 3

    catalog = parse_catalog()

    parse_items_and_save(f"main_ {dt.now()}.xlsx", catalog, brands)
    end_time = dt.now()

    logging.info(f"Программа отработала за: {(end_time - start_time).total_seconds()} сек.")
    logging.info("PROGRAM EXITED")
